from datetime import date, timedelta
from calendar import monthrange
from typing import Optional

from fastapi import FastAPI, Depends, HTTPException, Response, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from apscheduler.schedulers.background import BackgroundScheduler
import logging
import os

from database import engine, get_db, Base, SessionLocal
import models
import schemas
import auth
import push

logger = logging.getLogger("ledger.main")

Base.metadata.create_all(bind=engine)

app = FastAPI(title="Personal Ledger")

FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend")


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def month_str(d: date) -> str:
    return d.strftime("%Y-%m")


def due_date_for(year: int, month: int, due_day: int) -> date:
    last_day = monthrange(year, month)[1]
    return date(year, month, min(due_day, last_day))


def ensure_current_month_entries(db: Session, today: Optional[date] = None):
    """Make sure every active person has a payment row for the current month.
    Safe to call as often as you like — it's a no-op if the row exists."""
    today = today or date.today()
    cm = month_str(today)
    people = db.query(models.Person).filter(models.Person.archived == False).all()  # noqa: E712
    for p in people:
        exists = db.query(models.PersonPayment).filter_by(person_id=p.id, month=cm).first()
        if not exists:
            db.add(models.PersonPayment(
                person_id=p.id, month=cm,
                due_date=due_date_for(today.year, today.month, p.due_day),
                paid_amount=0, paid_date=None,
            ))
    db.commit()


def get_or_create_settings(db: Session) -> models.Settings:
    s = db.query(models.Settings).first()
    if not s:
        s = models.Settings(id=1, cash_balance=0, savings_goal=0, current_savings=0)
        db.add(s)
        db.commit()
        db.refresh(s)
    return s


def payment_status(paid_amount: float, monthly_due: float, due_date: date, today: date) -> str:
    if paid_amount >= monthly_due:
        return "paid"
    if paid_amount > 0:
        return "partial"
    days = (due_date - today).days
    if days < 0:
        return "overdue"
    if days <= 2:
        return "due_soon"
    return "pending"


def fy_bounds(today: date):
    """Indian financial year: Apr 1 - Mar 31."""
    if today.month >= 4:
        start = date(today.year, 4, 1)
        end = date(today.year + 1, 3, 31)
    else:
        start = date(today.year - 1, 4, 1)
        end = date(today.year, 3, 31)
    return start, end


# ---------------------------------------------------------------------------
# auth routes
# ---------------------------------------------------------------------------
@app.get("/api/auth/status")
def auth_status(db: Session = Depends(get_db)):
    settings = db.query(models.Settings).first()
    has_password = bool(settings and settings.password_hash)
    return {"password_set": has_password}


@app.post("/api/auth/set-password")
def set_password(body: schemas.SetPasswordRequest, response: Response, db: Session = Depends(get_db)):
    settings = get_or_create_settings(db)
    if settings.password_hash:
        raise HTTPException(400, "Password already set. Use /api/auth/login instead.")
    if len(body.password) < 4:
        raise HTTPException(400, "Password must be at least 4 characters.")
    settings.password_hash = auth.hash_password(body.password)
    db.commit()
    token = auth.create_session_token()
    response.set_cookie(auth.COOKIE_NAME, token, httponly=True, samesite="lax",
                         max_age=auth.SESSION_MAX_AGE)
    return {"ok": True}


@app.post("/api/auth/login")
def login(body: schemas.LoginRequest, response: Response, db: Session = Depends(get_db)):
    settings = db.query(models.Settings).first()
    if not settings or not settings.password_hash:
        raise HTTPException(400, "No password set yet.")
    if not auth.verify_password(body.password, settings.password_hash):
        raise HTTPException(401, "Incorrect password.")
    token = auth.create_session_token()
    response.set_cookie(auth.COOKIE_NAME, token, httponly=True, samesite="lax",
                         max_age=auth.SESSION_MAX_AGE)
    return {"ok": True}


@app.post("/api/auth/logout")
def logout(response: Response):
    response.delete_cookie(auth.COOKIE_NAME)
    return {"ok": True}


# ---------------------------------------------------------------------------
# people
# ---------------------------------------------------------------------------
@app.get("/api/people", response_model=list[schemas.PersonOut], dependencies=[Depends(auth.require_auth)])
def list_people(include_archived: bool = False, db: Session = Depends(get_db)):
    ensure_current_month_entries(db)
    q = db.query(models.Person)
    if not include_archived:
        q = q.filter(models.Person.archived == False)  # noqa: E712
    return q.order_by(models.Person.name).all()


@app.post("/api/people", response_model=schemas.PersonOut, dependencies=[Depends(auth.require_auth)])
def create_person(body: schemas.PersonCreate, db: Session = Depends(get_db)):
    if body.due_day < 1 or body.due_day > 31:
        raise HTTPException(400, "due_day must be between 1 and 31")
    person = models.Person(
        name=body.name, amount_given=body.amount_given,
        monthly_due=body.monthly_due, due_day=body.due_day,
    )
    db.add(person)
    db.commit()
    db.refresh(person)
    ensure_current_month_entries(db)
    db.refresh(person)
    return person


@app.get("/api/people/{person_id}", response_model=schemas.PersonOut, dependencies=[Depends(auth.require_auth)])
def get_person(person_id: int, db: Session = Depends(get_db)):
    person = db.query(models.Person).get(person_id)
    if not person:
        raise HTTPException(404, "Person not found")
    return person


@app.post("/api/people/{person_id}/archive", response_model=schemas.PersonOut, dependencies=[Depends(auth.require_auth)])
def toggle_archive(person_id: int, db: Session = Depends(get_db)):
    person = db.query(models.Person).get(person_id)
    if not person:
        raise HTTPException(404, "Person not found")
    person.archived = not person.archived
    db.commit()
    db.refresh(person)
    return person


@app.delete("/api/people/{person_id}", dependencies=[Depends(auth.require_auth)])
def delete_person(person_id: int, db: Session = Depends(get_db)):
    person = db.query(models.Person).get(person_id)
    if not person:
        raise HTTPException(404, "Person not found")
    db.delete(person)
    db.commit()
    return {"ok": True}


@app.post("/api/people/{person_id}/payments", response_model=schemas.PersonOut, dependencies=[Depends(auth.require_auth)])
def record_payment(person_id: int, body: schemas.RecordPaymentRequest, db: Session = Depends(get_db)):
    person = db.query(models.Person).get(person_id)
    if not person:
        raise HTTPException(404, "Person not found")
    if body.amount <= 0:
        raise HTTPException(400, "Amount must be positive")

    rec = db.query(models.PersonPayment).filter_by(person_id=person_id, month=body.month).first()
    if not rec:
        year, mo = (int(x) for x in body.month.split("-"))
        rec = models.PersonPayment(
            person_id=person_id, month=body.month,
            due_date=due_date_for(year, mo, person.due_day), paid_amount=0,
        )
        db.add(rec)

    rec.paid_amount += body.amount
    rec.paid_date = body.paid_date or date.today()
    if body.note:
        # append rather than overwrite, so a history of notes across partial
        # payments in the same month isn't lost
        rec.note = f"{rec.note}; {body.note}" if rec.note else body.note
    db.commit()
    db.refresh(person)
    return person


@app.post("/api/people/{person_id}/payments/complete", response_model=schemas.PersonOut, dependencies=[Depends(auth.require_auth)])
def complete_payment(person_id: int, body: schemas.CompletePaymentRequest, db: Session = Depends(get_db)):
    """Manually mark a month as fully paid, regardless of what's been logged
    so far — for when you've settled up in person and just want it to show
    Completed without walking through the amount math."""
    person = db.query(models.Person).get(person_id)
    if not person:
        raise HTTPException(404, "Person not found")

    rec = db.query(models.PersonPayment).filter_by(person_id=person_id, month=body.month).first()
    if not rec:
        year, mo = (int(x) for x in body.month.split("-"))
        rec = models.PersonPayment(
            person_id=person_id, month=body.month,
            due_date=due_date_for(year, mo, person.due_day), paid_amount=0,
        )
        db.add(rec)

    rec.paid_amount = person.monthly_due
    rec.paid_date = body.paid_date or date.today()
    if body.note:
        rec.note = f"{rec.note}; {body.note}" if rec.note else body.note
    db.commit()
    db.refresh(person)
    return person


@app.get("/api/people/{person_id}/export", dependencies=[Depends(auth.require_auth)])
def export_person_history(person_id: int, db: Session = Depends(get_db)):
    person = db.query(models.Person).get(person_id)
    if not person:
        raise HTTPException(404, "Person not found")

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Payment history"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B2140", end_color="1B2140", fill_type="solid")
    normal_font = Font(name="Arial")
    currency_fmt = '₹#,##0.00'

    # summary block
    ws["A1"] = "Name"
    ws["B1"] = person.name
    ws["A2"] = "Amount given"
    ws["B2"] = person.amount_given
    ws["B2"].number_format = currency_fmt
    ws["A3"] = "Monthly due"
    ws["B3"] = person.monthly_due
    ws["B3"].number_format = currency_fmt
    ws["A4"] = "Due day of month"
    ws["B4"] = person.due_day
    for cell in ["A1", "A2", "A3", "A4"]:
        ws[cell].font = Font(name="Arial", bold=True)
    for cell in ["B1", "B2", "B3", "B4"]:
        ws[cell].font = normal_font

    headers = ["Month", "Due date", "Paid amount", "Paid date", "Status", "Note"]
    header_row = 6
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    today = date.today()
    rows = sorted(person.payments, key=lambda h: h.month, reverse=True)
    for i, h in enumerate(rows, start=header_row + 1):
        status = payment_status(h.paid_amount, person.monthly_due, h.due_date, today)
        status_label = {"paid": "Completed", "partial": "Partial", "overdue": "Overdue",
                         "due_soon": "Due soon", "pending": "Pending"}.get(status, status)
        ws.cell(row=i, column=1, value=h.month).font = normal_font
        c = ws.cell(row=i, column=2, value=h.due_date)
        c.number_format = "DD/MM/YYYY"
        c.font = normal_font
        c = ws.cell(row=i, column=3, value=h.paid_amount)
        c.number_format = currency_fmt
        c.font = normal_font
        c = ws.cell(row=i, column=4, value=h.paid_date)
        if h.paid_date:
            c.number_format = "DD/MM/YYYY"
        c.font = normal_font
        ws.cell(row=i, column=5, value=status_label).font = normal_font
        ws.cell(row=i, column=6, value=h.note or "").font = normal_font

    widths = [12, 14, 14, 14, 12, 30]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = "".join(c for c in person.name if c.isalnum() or c in " _-").strip() or "person"
    headers = {"Content-Disposition": f'attachment; filename="{safe_name}-history.xlsx"'}
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# ---------------------------------------------------------------------------
# expenses
# ---------------------------------------------------------------------------
@app.get("/api/expenses", response_model=list[schemas.ExpenseOut], dependencies=[Depends(auth.require_auth)])
def list_expenses(month: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(models.Expense)
    if month:
        start = date(int(month[:4]), int(month[5:7]), 1)
        end_day = monthrange(start.year, start.month)[1]
        end = date(start.year, start.month, end_day)
        q = q.filter(models.Expense.date >= start, models.Expense.date <= end)
    return q.order_by(models.Expense.date.desc()).all()


@app.post("/api/expenses", response_model=schemas.ExpenseOut, dependencies=[Depends(auth.require_auth)])
def create_expense(body: schemas.ExpenseCreate, db: Session = Depends(get_db)):
    expense = models.Expense(**body.model_dump())
    db.add(expense)
    db.commit()
    db.refresh(expense)
    return expense


@app.delete("/api/expenses/{expense_id}", dependencies=[Depends(auth.require_auth)])
def delete_expense(expense_id: int, db: Session = Depends(get_db)):
    expense = db.query(models.Expense).get(expense_id)
    if not expense:
        raise HTTPException(404, "Expense not found")
    db.delete(expense)
    db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# expense categories
# ---------------------------------------------------------------------------
DEFAULT_CATEGORIES = ["Rent", "Groceries", "Travel", "Utilities", "Family Support", "Eating Out", "Other"]


def ensure_default_categories(db: Session):
    if db.query(models.ExpenseCategory).count() == 0:
        for name in DEFAULT_CATEGORIES:
            db.add(models.ExpenseCategory(name=name))
        db.commit()


@app.get("/api/expense-categories", response_model=list[schemas.ExpenseCategoryOut], dependencies=[Depends(auth.require_auth)])
def list_expense_categories(db: Session = Depends(get_db)):
    ensure_default_categories(db)
    return db.query(models.ExpenseCategory).order_by(models.ExpenseCategory.name).all()


@app.post("/api/expense-categories", response_model=schemas.ExpenseCategoryOut, dependencies=[Depends(auth.require_auth)])
def create_expense_category(body: schemas.ExpenseCategoryCreate, db: Session = Depends(get_db)):
    name = body.name.strip()
    if not name:
        raise HTTPException(400, "Category name can't be empty")
    existing = db.query(models.ExpenseCategory).filter(func.lower(models.ExpenseCategory.name) == name.lower()).first()
    if existing:
        return existing
    cat = models.ExpenseCategory(name=name)
    db.add(cat)
    db.commit()
    db.refresh(cat)
    return cat


@app.delete("/api/expense-categories/{category_id}", dependencies=[Depends(auth.require_auth)])
def delete_expense_category(category_id: int, db: Session = Depends(get_db)):
    cat = db.query(models.ExpenseCategory).get(category_id)
    if not cat:
        raise HTTPException(404, "Category not found")
    db.delete(cat)
    db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# savings log
# ---------------------------------------------------------------------------
@app.get("/api/savings-log", response_model=list[schemas.SavingsEntryOut], dependencies=[Depends(auth.require_auth)])
def list_savings_log(db: Session = Depends(get_db)):
    return db.query(models.SavingsEntry).order_by(models.SavingsEntry.date.desc()).all()


@app.post("/api/savings-log", response_model=schemas.SavingsEntryOut, dependencies=[Depends(auth.require_auth)])
def create_savings_entry(body: schemas.SavingsEntryCreate, db: Session = Depends(get_db)):
    entry = models.SavingsEntry(**body.model_dump())
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return entry


@app.delete("/api/savings-log/{entry_id}", dependencies=[Depends(auth.require_auth)])
def delete_savings_entry(entry_id: int, db: Session = Depends(get_db)):
    entry = db.query(models.SavingsEntry).get(entry_id)
    if not entry:
        raise HTTPException(404, "Entry not found")
    db.delete(entry)
    db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# settings
# ---------------------------------------------------------------------------
@app.get("/api/settings", response_model=schemas.SettingsOut, dependencies=[Depends(auth.require_auth)])
def get_settings(db: Session = Depends(get_db)):
    s = get_or_create_settings(db)
    return s


@app.patch("/api/settings", response_model=schemas.SettingsOut, dependencies=[Depends(auth.require_auth)])
def update_settings(body: schemas.SettingsUpdate, db: Session = Depends(get_db)):
    s = get_or_create_settings(db)
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(s, field, value)
    db.commit()
    db.refresh(s)
    return s


# ---------------------------------------------------------------------------
# push notifications
# ---------------------------------------------------------------------------
@app.get("/api/push/public-key")
def push_public_key():
    """No auth required — the browser needs this before you're necessarily
    logged in yet, and it's not sensitive (it's meant to be public)."""
    return {"publicKey": push.get_or_create_vapid_keys()["public_key"]}


@app.post("/api/push/subscribe", dependencies=[Depends(auth.require_auth)])
def push_subscribe(body: schemas.PushSubscribeRequest, db: Session = Depends(get_db)):
    existing = db.query(models.PushSubscription).filter_by(endpoint=body.endpoint).first()
    if existing:
        existing.p256dh = body.keys.p256dh
        existing.auth = body.keys.auth
    else:
        db.add(models.PushSubscription(
            endpoint=body.endpoint, p256dh=body.keys.p256dh, auth=body.keys.auth,
        ))
    db.commit()
    return {"ok": True}


@app.post("/api/push/unsubscribe", dependencies=[Depends(auth.require_auth)])
def push_unsubscribe(body: schemas.PushUnsubscribeRequest, db: Session = Depends(get_db)):
    db.query(models.PushSubscription).filter_by(endpoint=body.endpoint).delete()
    db.commit()
    return {"ok": True}


@app.post("/api/push/test", dependencies=[Depends(auth.require_auth)])
def push_test(db: Session = Depends(get_db)):
    subs = db.query(models.PushSubscription).all()
    if not subs:
        raise HTTPException(400, "No devices are subscribed yet. Tap 'Enable notifications' first.")
    sent = 0
    for s in subs:
        ok = push.send_notification(
            {"endpoint": s.endpoint, "keys": {"p256dh": s.p256dh, "auth": s.auth}},
            "Ledger test notification",
            "If you can see this, notifications are working.",
        )
        if ok:
            sent += 1
        else:
            # subscription is likely dead (expired/revoked) — remove it
            db.delete(s)
    db.commit()
    return {"sent": sent, "total": len(subs)}


def run_due_date_check():
    """Runs once a day (see scheduler setup below). Sends one digest
    notification per subscribed device listing anyone overdue or due
    within 2 days. Safe to call manually too — it's idempotent per call,
    though it will re-notify for the same unpaid month on each run since
    there's no separate 'already notified' tracking — intentionally simple."""
    db = SessionLocal()
    try:
        today = date.today()
        cm = month_str(today)
        ensure_current_month_entries(db, today)
        people = db.query(models.Person).filter(models.Person.archived == False).all()  # noqa: E712

        overdue_names, due_soon_names = [], []
        for p in people:
            rec = next((pay for pay in p.payments if pay.month == cm), None)
            if not rec:
                continue
            status = payment_status(rec.paid_amount, p.monthly_due, rec.due_date, today)
            if status == "overdue":
                overdue_names.append(p.name)
            elif status == "due_soon":
                due_soon_names.append(p.name)

        if not overdue_names and not due_soon_names:
            return

        parts = []
        if overdue_names:
            parts.append(f"Overdue: {', '.join(overdue_names)}")
        if due_soon_names:
            parts.append(f"Due soon: {', '.join(due_soon_names)}")
        body = " · ".join(parts)

        subs = db.query(models.PushSubscription).all()
        for s in subs:
            ok = push.send_notification(
                {"endpoint": s.endpoint, "keys": {"p256dh": s.p256dh, "auth": s.auth}},
                "Payments need attention", body,
            )
            if not ok:
                db.delete(s)
        db.commit()
    except Exception:
        logger.exception("Daily due-date check failed")
    finally:
        db.close()


_scheduler = BackgroundScheduler()


@app.on_event("startup")
def start_scheduler():
    notify_hour = int(os.getenv("LEDGER_NOTIFY_HOUR", "9"))
    _scheduler.add_job(run_due_date_check, "cron", hour=notify_hour, minute=0, id="due_date_check")
    _scheduler.start()
    logger.info("Scheduled daily due-date check at %02d:00 local time", notify_hour)


@app.on_event("shutdown")
def stop_scheduler():
    _scheduler.shutdown(wait=False)


# ---------------------------------------------------------------------------
# dashboard / overview
# ---------------------------------------------------------------------------
@app.get("/api/dashboard", dependencies=[Depends(auth.require_auth)])
def dashboard(db: Session = Depends(get_db)):
    today = date.today()
    cm = month_str(today)
    ensure_current_month_entries(db, today)

    people = db.query(models.Person).filter(models.Person.archived == False).all()  # noqa: E712

    notifications = {"overdue": [], "due_soon": [], "pending": []}
    month_collected = 0.0
    total_outstanding = 0.0
    total_principal_out = 0.0

    for p in people:
        cm_rec = next((pay for pay in p.payments if pay.month == cm), None)
        collected_all_time = sum(pay.paid_amount for pay in p.payments)
        expected_so_far = sum(p.monthly_due for pay in p.payments if pay.month <= cm)
        outstanding = max(0.0, expected_so_far - collected_all_time)
        principal_outstanding = max(0.0, p.amount_given - collected_all_time)

        total_outstanding += outstanding
        total_principal_out += principal_outstanding

        if cm_rec:
            month_collected += cm_rec.paid_amount
            status = payment_status(cm_rec.paid_amount, p.monthly_due, cm_rec.due_date, today)
            entry = {
                "person_id": p.id, "name": p.name, "monthly_due": p.monthly_due,
                "due_day": p.due_day, "due_date": cm_rec.due_date.isoformat(),
                "paid_amount": cm_rec.paid_amount, "status": status,
            }
            if status == "overdue":
                notifications["overdue"].append(entry)
            elif status == "due_soon":
                notifications["due_soon"].append(entry)
            elif status == "pending":
                notifications["pending"].append(entry)

    month_start = date(today.year, today.month, 1)
    month_end = date(today.year, today.month, monthrange(today.year, today.month)[1])
    month_expenses = db.query(func.sum(models.Expense.amount)).filter(
        models.Expense.date >= month_start, models.Expense.date <= month_end
    ).scalar() or 0.0

    settings = get_or_create_settings(db)
    net_worth = settings.cash_balance + total_principal_out

    return {
        "month": cm,
        "month_collected": month_collected,
        "total_outstanding": total_outstanding,
        "month_expenses": month_expenses,
        "notifications": notifications,
        "net_worth": {
            "cash_balance": settings.cash_balance,
            "principal_out": total_principal_out,
            "net_worth": net_worth,
        },
        "savings_goal": {
            "current": settings.current_savings,
            "goal": settings.savings_goal,
            "pct": round(100 * settings.current_savings / settings.savings_goal, 1) if settings.savings_goal else 0,
        },
        "people_status": [
            {
                "person_id": p.id, "name": p.name,
                "status": payment_status(
                    next((pay.paid_amount for pay in p.payments if pay.month == cm), 0),
                    p.monthly_due,
                    next((pay.due_date for pay in p.payments if pay.month == cm), today),
                    today,
                ),
                "paid_amount": next((pay.paid_amount for pay in p.payments if pay.month == cm), 0),
                "monthly_due": p.monthly_due,
            }
            for p in people
        ],
    }


# ---------------------------------------------------------------------------
# summary (financial year)
# ---------------------------------------------------------------------------
@app.get("/api/summary", dependencies=[Depends(auth.require_auth)])
def summary(db: Session = Depends(get_db)):
    today = date.today()
    fy_start, fy_end = fy_bounds(today)

    expenses = db.query(models.Expense).filter(
        models.Expense.date >= fy_start, models.Expense.date <= fy_end
    ).all()
    total_expenses = sum(e.amount for e in expenses)

    by_category = {}
    for e in expenses:
        by_category[e.category] = by_category.get(e.category, 0) + e.amount
    by_category_sorted = sorted(by_category.items(), key=lambda kv: -kv[1])

    all_people = db.query(models.Person).all()
    total_collected = 0.0
    for p in all_people:
        for pay in p.payments:
            y, m = (int(x) for x in pay.month.split("-"))
            pay_date = date(y, m, 1)
            if fy_start <= pay_date <= fy_end:
                total_collected += pay.paid_amount

    return {
        "fy_start": fy_start.isoformat(),
        "fy_end": fy_end.isoformat(),
        "total_collected": total_collected,
        "total_expenses": total_expenses,
        "by_category": [{"category": c, "amount": a} for c, a in by_category_sorted],
    }


# ---------------------------------------------------------------------------
# backup / export
# ---------------------------------------------------------------------------
@app.get("/api/backup", dependencies=[Depends(auth.require_auth)])
def backup(db: Session = Depends(get_db)):
    people = db.query(models.Person).all()
    expenses = db.query(models.Expense).all()
    settings = get_or_create_settings(db)

    data = {
        "exported_at": date.today().isoformat(),
        "people": [
            {
                "name": p.name, "amount_given": p.amount_given, "monthly_due": p.monthly_due,
                "due_day": p.due_day, "archived": p.archived,
                "payments": [
                    {"month": pay.month, "due_date": pay.due_date.isoformat(),
                     "paid_amount": pay.paid_amount,
                     "paid_date": pay.paid_date.isoformat() if pay.paid_date else None}
                    for pay in p.payments
                ],
            }
            for p in people
        ],
        "expenses": [
            {"date": e.date.isoformat(), "category": e.category, "amount": e.amount, "note": e.note}
            for e in expenses
        ],
        "settings": {
            "cash_balance": settings.cash_balance,
            "savings_goal": settings.savings_goal,
            "current_savings": settings.current_savings,
        },
    }
    headers = {"Content-Disposition": f'attachment; filename="ledger-backup-{date.today().isoformat()}.json"'}
    return JSONResponse(content=data, headers=headers)


# ---------------------------------------------------------------------------
# static frontend
# ---------------------------------------------------------------------------
app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")


@app.get("/")
def serve_index():
    return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))
